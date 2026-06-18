import os
import re
import json
import tempfile
import shutil
import requests
import traceback
from dotenv import load_dotenv
from flask import Flask, request, jsonify
import gdown

load_dotenv()
api = Flask(__name__)

# ========== 1. НАСТРОЙКИ ==========
BITRIX24_WEBHOOK = os.getenv("BITRIX24_WEBHOOK")
KODIK_API_KEY = os.getenv("KODIK_API_KEY")
DEAL_CATEGORY_ID = int(os.getenv("DEAL_CATEGORY_ID", 42))
DEAL_STAGE_ID = os.getenv("DEAL_STAGE_ID", "8704")
FIELD_LINK_CODE = os.getenv("FIELD_LINK_CODE", "UF_CRM_1774428455758")
FIELD_COMPANY_DIRECTION = os.getenv("FIELD_COMPANY_DIRECTION", "UF_CRM_1774954195201")
FIELD_DRIVE_FOLDER_LINK = "UF_CRM_1781350841203"
GOOGLE_DRIVE_ROOT_FOLDER_ID = os.getenv("GOOGLE_DRIVE_ROOT_FOLDER_ID")
DRIVE_API_KEY = os.getenv("DRIVE_API_KEY")

DEEPSEEK_URL = "https://api.kodikrouter.ru/v1/chat/completions"
DEEPSEEK_MODEL = "deepseek/deepseek-v4-pro"

PROMPT = """
Во вложении техническая документация по закупке.
Проанализируй документы и предоставь подробную информацию в отчет удобный для копирования в документ WORD:
1) Название кампании Заказчика, его ИНН и контакты сотрудников если такие представлены (с указанием номера телефона, почты, должности и ФИО). Отдельно выпиши ФИО представителя Заказчика подготовившего ТЗ. Так же кратко укажи чем занимается организация ее отрасль;
2) Адреса проведения работ;
3) Название закупки и обоснование для проведения работ, если такая информация есть;
3.1 Есть ли общий выделенный бюджет на закупку? в смете или в НМЦ (Если в НМЦ возьми самую минимальную цену предложенную в КП)
4) Перечисли указанные в документации: 
4.1) Перечень оборудования (Например: котел/теплообменник/калорифер/реактор и т.д.), его наименование, его модель (Например: Visman vitomasx 200-WS или Alfa Laval	A15BW), количество оборудования, вид работ с этим оборудованием (Например: химическая промывка/механическая промывка/разборная промывка/Без разборная промывка),

Важно! Всегда сначала проверь есть ли во вложенной документации необходимая информация, в случае если информация присутствует, напиши ее, если информации нет, то проверить ее наличие в открытом доступе интернет, если найти ее удалось со 100% точностью, то напиши результат с пометкой "из открытых источников", если найти в открытых источниках не удалось, то напиши "уточнить у Заказчика".

Важно! Если работы предусматривают промывку котла/реактора/емкости или др. сосудов: необходимо указать объем водяной рубашки для этого оборудования.

Важно! Если работы предусматривают промывку теплообменников, то прежде всего необходимо определить и указать его тип: пластинчатый, паянный кожухотрубный т.д., 
Далее необходимо определить и указать вид промывки: для пластинчатых теплообменников это может быть разборная, без разборная, механическая промывка и другая указанная в документации. Для Паянных аппаратов, только без разборная, химическая если в Документации не указано иное. Для кожухотрубных может быть механическая, без разборная промывка. 
Если промывка Разборная, то обязательно указать количество пластин в каждом теплообменнике их размеры пластин (высота и ширина), размер Ду(DN). Если работы предусматривают промывку теплообменников пластинчатых или паянных безразборно, то обязательно указать размер ДУ(Dn) присоединений и размеры пластин(высота и ширина). Если работы предусматривают промывку теплообменника кожухотрубного, то необходимо выяснить и указать объем этого теплообменника, количество трубок и их диаметр Ду(DN).

Важно!! Если в документах Заказчика нет информации о размерах, объемах оборудования, то всегда сверься с приложенным файлом "Реестр оборудования и его размеров.xlsx" приложенным к запросу, информация из документов (ТЗ, паспорта) является приоритетной, обязательно напиши в таблице если взял данные из нашего реестра.
Результат по пункту 4.1 выведи в формате таблицы.

4.2 Укажи Перечень дополнительных требований к выполнению работ(Например: ультразвук, гидроипульсы, баражирование и т.д).
4.3 Укажи требования к результатам работ и способ которым это будет фиксироваться;
5) Укажи есть ли требования к подбору моющего средства (Например: требуется ли определенное средство, биоорганическое, кислотное, либо наоборот без содержания соляной кислоты и т.д.)
Обязательно укажи требования по утилизации объема отработанного раствора/моющего средства.

6) Укажи есть ли требования к персоналу? (Например: наличие определенных допусков к работе, определенное количество сотрудников и требования к их квалификации)

7) Укажи есть ли требования к опыту кампании, нужно ли его подтвердить. Как? Например: наличие договор на оказание услуг суммой 3 000 000 рублей миниму за последний год.
7) Укажи сроки проведения работ

8) Укажи условия оплаты

9) Укажи порядок определения победителя

10) Предусмотренные штрафы, санкции, неустойки. Цена? за что?

11) Ниже выпиши ключевые вопросы которые необходимо уточнить у Заказчика
"""

# ========== 2. ФУНКЦИИ БИТРИКС24 ==========
def call_bitrix24(method, params):
    url = f"{BITRIX24_WEBHOOK}{method}"
    try:
        resp = requests.post(url, json=params, timeout=60)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        return {"error": str(e)}

def get_enum_value_id(field_name, target_value):
    fields_response = call_bitrix24("crm.company.fields", {})
    if "error" in fields_response:
        raise Exception(f"Не удалось получить поля компаний: {fields_response['error']}")
    field_info = fields_response["result"].get(field_name)
    if not field_info:
        raise Exception(f"Поле {field_name} не найдено")
    for item in field_info.get("items", []):
        if item.get("VALUE") == target_value:
            return item.get("ID")
    raise Exception(f"Значение '{target_value}' не найдено для поля {field_name}")

def find_company_by_inn(inn):
    print(f"[LOG] Поиск компании по ИНН: '{inn}' (длина {len(inn) if inn else 0})")
    if not inn or inn.strip() == '':
        print("[LOG] ИНН пустой, возвращаем None")
        return None
    inn_clean = inn.strip()
    print(f"[LOG] ИНН после очистки: '{inn_clean}' (длина {len(inn_clean)})")
    filter_params = {"RQ_INN": inn_clean, "ENTITY_TYPE_ID": 4}
    print(f"[LOG] Фильтр para crm.requisite.list: {json.dumps(filter_params, ensure_ascii=False)}")
    result = call_bitrix24("crm.requisite.list", {
        "filter": filter_params,
        "select": ["ID", "ENTITY_ID"]
    })
    print(f"[LOG] Ответ от crm.requisite.list (первые 500 символов): {json.dumps(result, ensure_ascii=False)[:500]}")
    if result.get("error"):
        print(f"[LOG] Ошибка поиска реквизитов: {result['error']}")
        return None
    if result.get("result"):
        requisite = result["result"][0]
        company_id = requisite.get("ENTITY_ID")
        print(f"[LOG] Найдена компания с ID {company_id} для ИНН {inn_clean}")
        return company_id
    else:
        print("[LOG] Реквизиты с таким ИНН не найдены")
        return None

def create_company(company_name, inn):
    direction_id = get_enum_value_id(FIELD_COMPANY_DIRECTION, "НВ")
    company_res = call_bitrix24("crm.company.add", {
        "fields": {
            "TITLE": company_name,
            FIELD_COMPANY_DIRECTION: direction_id
        }
    })
    company_id = company_res.get("result")
    if not company_id:
        raise Exception(f"Ошибка создания компании: {company_res}")
    presets = call_bitrix24("crm.requisite.preset.list", {})
    preset_id = 1
    if presets.get("result"):
        for p in presets["result"]:
            if "юр" in p.get("NAME", "").lower() or "organiz" in p.get("NAME", "").lower():
                preset_id = p["ID"]
                break
    call_bitrix24("crm.requisite.add", {
        "fields": {
            "ENTITY_TYPE_ID": 4,
            "ENTITY_ID": company_id,
            "PRESET_ID": preset_id,
            "NAME": company_name,
            "RQ_INN": inn
        }
    })
    return company_id

def find_contact_by_email_or_phone(email, phone):
    """Ищет контакт в Битрикс24 по email или телефону. Возвращает ID или None."""
    if not email and not phone:
        return None
    
    # Сначала пробуем найти по email (точное совпадение)
    if email:
        result = call_bitrix24("crm.contact.list", {
            "filter": {"EMAIL": email},
            "select": ["ID"]
        })
        if result.get("result") and len(result["result"]) > 0:
            contact_id = result["result"][0]["ID"]
            print(f"[LOG] Найден существующий контакт по email {email}: ID {contact_id}")
            return contact_id
    
    # Если email не найден или нет email, пробуем по телефону (точное совпадение)
    if phone:
        result = call_bitrix24("crm.contact.list", {
            "filter": {"PHONE": phone},
            "select": ["ID"]
        })
        if result.get("result") and len(result["result"]) > 0:
            contact_id = result["result"][0]["ID"]
            print(f"[LOG] Найден существующий контакт по телефону {phone}: ID {contact_id}")
            return contact_id
    
    print("[LOG] Контакт не найден, будем создавать новый")
    return None

def create_contact(name, phone, email, company_id):
    if not any([name, phone, email]):
        return None
    
    # Проверяем, есть ли уже такой контакт
    existing_contact_id = find_contact_by_email_or_phone(email, phone)
    if existing_contact_id:
        # Если контакт уже существует, просто привязываем его к компании (если ещё не привязан)
        print(f"[LOG] Используем существующий контакт ID {existing_contact_id}")
        # Проверяем, привязан ли контакт к компании
        check = call_bitrix24("crm.company.contact.get", {"id": company_id})
        if check.get("result"):
            already_linked = any(c["CONTACT_ID"] == existing_contact_id for c in check["result"])
            if not already_linked:
                call_bitrix24("crm.company.contact.add", {
                    "id": company_id,
                    "fields": {"CONTACT_ID": existing_contact_id}
                })
                print(f"[LOG] Контакт {existing_contact_id} привязан к компании {company_id}")
        return existing_contact_id
    
    # Если контакт не найден, создаём новый
    contact_res = call_bitrix24("crm.contact.add", {
        "fields": {
            "NAME": name or "",
            "PHONE": [{"VALUE": phone, "VALUE_TYPE": "WORK"}] if phone else [],
            "EMAIL": [{"VALUE": email, "VALUE_TYPE": "WORK"}] if email else []
        }
    })
    contact_id = contact_res.get("result")
    if contact_id and company_id:
        call_bitrix24("crm.company.contact.add", {
            "id": company_id,
            "fields": {"CONTACT_ID": contact_id}
        })
        print(f"[LOG] Создан новый контакт ID {contact_id} и привязан к компании {company_id}")
    return contact_id

def create_deal(company_id, deal_name, purchase_link, drive_folder_link):
    deal_fields = {
        "TITLE": deal_name,
        "COMPANY_ID": company_id,
        "CATEGORY_ID": DEAL_CATEGORY_ID,
        "STAGE_ID": DEAL_STAGE_ID,
        FIELD_LINK_CODE: purchase_link,
        FIELD_DRIVE_FOLDER_LINK: drive_folder_link
    }
    deal_res = call_bitrix24("crm.deal.add", {"fields": deal_fields})
    return deal_res.get("result")

def markdown_table_to_bbcode(table_lines):
    if len(table_lines) < 2:
        return '\n'.join(table_lines)
    header_line = table_lines[0].strip('|').split('|')
    headers = [h.strip() for h in header_line]
    data_rows = []
    for line in table_lines[2:]:
        cells = line.strip('|').split('|')
        row = [c.strip() for c in cells]
        data_rows.append(row)
    bbcode = '[table]\n'
    bbcode += '[tr]' + ''.join(f'[th]{h}[/th]' for h in headers) + '[/tr]\n'
    for row in data_rows:
        bbcode += '[tr]' + ''.join(f'[td]{cell}[/td]' for cell in row) + '[/tr]\n'
    bbcode += '[/table]'
    return bbcode

def markdown_to_bbcode(text):
    lines = text.split('\n')
    result = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.strip().startswith('|') and i+1 < len(lines) and re.match(r'^[\s]*\|[\s\-:]+[\s]*\|', lines[i+1]):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                table_lines.append(lines[i].strip())
                i += 1
            result.append(markdown_table_to_bbcode(table_lines))
            continue
        else:
            line = re.sub(r'\*\*(.*?)\*\*', r'[b]\1[/b]', line)
            line = re.sub(r'__(.*?)__', r'[b]\1[/b]', line)
            line = re.sub(r'\*(.*?)\*', r'[i]\1[/i]', line)
            line = re.sub(r'_(.*?)_', r'[i]\1[/i]', line)
            line = re.sub(r'^# (.*?)$', r'[size=18][b]\1[/b][/size]', line, flags=re.MULTILINE)
            line = re.sub(r'^## (.*?)$', r'[size=16][b]\1[/b][/size]', line, flags=re.MULTILINE)
            line = re.sub(r'^### (.*?)$', r'[size=14][b]\1[/b][/size]', line, flags=re.MULTILINE)
            if re.match(r'^[\*\-\+]\s+', line):
                line = re.sub(r'^[\*\-\+]\s+', '[*]', line)
            result.append(line)
            i += 1
    full_text = '\n'.join(result)
    full_text = re.sub(r'([\n]*)(\[\*].*?)(?=\n\[\*]|\Z)', r'\n[list]\2[/list]\n', full_text, flags=re.DOTALL)
    full_text = re.sub(r'\n{3,}', '\n\n', full_text)
    return full_text

def add_comment_to_deal(deal_id, comment_text):
    formatted_comment = markdown_to_bbcode(comment_text)
    params = {"fields": {"ENTITY_ID": deal_id, "ENTITY_TYPE": "deal", "COMMENT": formatted_comment}}
    return call_bitrix24("crm.timeline.comment.add", params)

# ========== 3. ФУНКЦИИ ДЛЯ GOOGLE DRIVE ==========
def find_subfolder_id_by_api(parent_folder_id, target_name):
    if not DRIVE_API_KEY:
        print("Ошибка: не задан DRIVE_API_KEY")
        return None
    url = "https://www.googleapis.com/drive/v3/files"
    query = f"'{parent_folder_id}' in parents and name='{target_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    params = {
        'q': query,
        'key': DRIVE_API_KEY,
        'fields': 'files(id, name)'
    }
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        files = data.get('files', [])
        if files:
            folder_id = files[0]['id']
            print(f"Найдена подпапка '{target_name}' с ID {folder_id}")
            return folder_id
        else:
            print(f"Подпапка с именем '{target_name}' не найдена.")
            return None
    except Exception as e:
        print(f"Ошибка при запросе к Google Drive API: {e}")
        return None

def download_folder_by_id(folder_id, destination_dir):
    try:
        print(f"Скачивание папки с ID {folder_id} через gdown...")
        gdown.download_folder(id=folder_id, output=destination_dir, use_cookies=False, quiet=False)
        downloaded_files = []
        for root, dirs, files in os.walk(destination_dir):
            for file in files:
                if not file.endswith('.html') and not file.startswith('.'):
                    downloaded_files.append(os.path.join(root, file))
        print(f"Скачано файлов: {len(downloaded_files)}")
        return downloaded_files
    except Exception as e:
        print(f"Ошибка скачивания папки {folder_id}: {e}")
        return []

def extract_text_from_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    print(f"Попытка извлечь текст из: {file_path} (расширение {ext})")
    try:
        if ext == '.pdf':
            import PyPDF2
            text = ""
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            return text
        elif ext == '.docx':
            import docx
            doc = docx.Document(file_path)
            return "\n".join([p.text for p in doc.paragraphs])
        elif ext == '.doc':
            # Пробуем прочитать как текст (бинарный файл даст мусор, но это лучше, чем ничего)
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                if len(content) > 50:
                    return content
                else:
                    print(f"Файл {file_path} содержит мало текста, возможно бинарный")
                    return None
        elif ext == '.txt':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        elif ext == '.xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text += " ".join(str(cell) for cell in row if cell) + "\n"
            return text
        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(file_path)
            text = ""
            for sheet in wb.sheets():
                for row in range(sheet.nrows):
                    row_text = []
                    for col in range(sheet.ncols):
                        cell = sheet.cell(row, col)
                        if cell.ctype != xlrd.XL_CELL_EMPTY:
                            row_text.append(str(cell.value))
                    text += " ".join(row_text) + "\n"
            return text
        else:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
    except Exception as e:
        print(f"Ошибка извлечения текста из {file_path}: {e}")
        return None

# ========== 4. ФУНКЦИЯ ДЛЯ DEEPSEEK ==========
def analyze_with_deepseek(file_paths):
    import time
    start_time = time.time()
    
    full_text = ""
    for path in file_paths:
        text = extract_text_from_file(path)
        if text is None:
            print(f"Файл {os.path.basename(path)} пропущен (не удалось извлечь текст)")
            continue
        if text:
            full_text += f"\n\n--- Файл: {os.path.basename(path)} ---\n{text}\n"
    if not full_text.strip():
        raise Exception("Не удалось извлечь текст ни из одного файла")
    if len(full_text) > 800000:
        full_text = full_text[:800000] + "...\n[Текст документа обрезан]"

    user_prompt = f"{PROMPT}\n\nТекст документов:\n{full_text}"
    headers = {
        "Authorization": f"Bearer {KODIK_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": DEEPSEEK_MODEL,
        "messages": [{"role": "user", "content": user_prompt}],
        "temperature": 0.3,
        "max_tokens": 32000  # <--- Увеличили с 8000 до 32000
    }
    try:
        print("Отправка запроса к KodikRouter...")
        response = requests.post(DEEPSEEK_URL, headers=headers, json=payload, timeout=300)
        print(f"Статус ответа KodikRouter: {response.status_code}")
        response.raise_for_status()
        data = response.json()
        
        # Логируем finish_reason и длину ответа
        if "choices" in data and len(data["choices"]) > 0:
            finish_reason = data["choices"][0].get("finish_reason")
            print(f"finish_reason: {finish_reason}")
        
        content = None
        if isinstance(data, dict):
            if "choices" in data and len(data["choices"]) > 0:
                msg = data["choices"][0].get("message")
                if msg and isinstance(msg, dict):
                    content = msg.get("content")
            if not content and "content" in data:
                content = data["content"]
            if not content and "response" in data:
                content = data["response"]
        elif isinstance(data, str):
            content = data
        
        elapsed = time.time() - start_time
        print(f"Время обработки запроса: {elapsed:.2f} сек")
        
        if content and isinstance(content, str):
            return content
        else:
            print(f"Не удалось извлечь content. Тип data: {type(data)}")
            print(f"Содержимое data: {json.dumps(data, ensure_ascii=False)[:500]}")
            return "Ошибка: не удалось получить ответ от модели"
    except Exception as e:
        print(f"Ошибка DeepSeek: {e}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"Тело ошибки: {e.response.text[:500]}")
        return f"Ошибка при обращении к KodikRouter: {str(e)}"

# ========== 5. ОСНОВНАЯ ЛОГИКА ==========
def process_purchase(data):
    try:
        print(f"[PROCESS] Начало обработки закупки {data.get('purchase_number', 'N/A')}")
        company_id = find_company_by_inn(data["inn"])
        if not company_id:
            company_id = create_company(data["company_name"], data["inn"])
            print(f"Создана новая компания ID {company_id}")
        else:
            print(f"Компания уже существует, ID {company_id}")

        contact_id = create_contact(data.get("contact_name"), data.get("phone"), data.get("email"), company_id)
        if contact_id:
            print(f"Создан контакт ID {contact_id} и привязан к компании {company_id}")

        purchase_number = data["purchase_number"]
        print(f"Поиск папки для закупки {purchase_number}...")
        subfolder_id = find_subfolder_id_by_api(GOOGLE_DRIVE_ROOT_FOLDER_ID, purchase_number)
        if not subfolder_id:
            return {"status": "error", "message": f"Не найдена подпапка с именем {purchase_number}"}

        drive_folder_link = f"https://drive.google.com/drive/folders/{subfolder_id}"
        deal_id = create_deal(company_id, data["company_name"], data.get("purchase_link", ""), drive_folder_link)
        print(f"Создана сделка ID {deal_id}")

        if contact_id:
            try:
                call_bitrix24("crm.deal.contact.add", {"id": deal_id, "fields": {"CONTACT_ID": contact_id}})
                print(f"Контакт {contact_id} привязан к сделке {deal_id}")
            except Exception as e:
                print(f"Ошибка при привязке контакта: {e}")

        print("[PROCESS] Начинаем скачивание файлов...")
        temp_dir = tempfile.mkdtemp()
        try:
            downloaded_files = download_folder_by_id(subfolder_id, temp_dir)
            if not downloaded_files:
                return {"status": "error", "message": f"Папка {purchase_number} пуста или не содержит поддерживаемых файлов. Проверьте, что файлы загружены."}

            print(f"Скачано файлов: {len(downloaded_files)}")
            for f in downloaded_files:
                print(f"  - {os.path.basename(f)}")

            print("[PROCESS] Отправка в DeepSeek...")
            analysis = analyze_with_deepseek(downloaded_files)
            if not isinstance(analysis, str):
                analysis = str(analysis)
            if not analysis:
                analysis = "Анализ не получен"

            comment_text = f"🤖 Анализ от DeepSeek (модель {DEEPSEEK_MODEL}):\n\n{analysis}"
            add_comment_to_deal(deal_id, comment_text)
            print("[PROCESS] Комментарий добавлен в сделку")

            analysis_preview = analysis[:300] if analysis and len(analysis) > 300 else (analysis or "Нет текста")
            return {"status": "success", "deal_id": deal_id, "analysis_preview": analysis_preview}
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception as e:
        print(f"[PROCESS] КРИТИЧЕСКАЯ ОШИБКА: {e}")
        traceback.print_exc()
        return {"status": "error", "message": str(e)}

# ========== 6. FLASK-ЭНДПОИНТ ==========
@api.route('/process', methods=['POST'])
def process_webhook():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "Нет данных"}), 400
        required = ["inn", "company_name", "purchase_number"]
        for field in required:
            if field not in data:
                return jsonify({"status": "error", "message": f"Отсутствует поле '{field}'"}), 400
        result = process_purchase(data)
        print(f"[FLASK] Ответ: {json.dumps(result, ensure_ascii=False)[:500]}")
        return jsonify(result)
    except Exception as e:
        print(f"[FLASK] Ошибка: {e}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

# ========== 7. ЗАПУСК ==========
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"[SERVER] Запуск на порту {port}")
    api.run(host='0.0.0.0', port=port)
